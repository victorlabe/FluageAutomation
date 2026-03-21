#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Fluage IUT d'Évreux — Application

Cette application de bureau (Tkinter + ttkbootstrap) permet d'acquérir et visualiser les données de fluage en temps réel depuis un Arduino Seeeduino connecté en USB.
Elle vient remplacer la méthode avec Excel PLX-DAQ, offrant une interface plus moderne, des graphiques intégrés et des fonctionnalités d'analyse avancées :
(calcul de la vitesse de déformation via régression linéaire).

------------------------------------------------------------------------------------------------------------

Fonctions principales
- Connexion à un port série (Arduino Seeeduino)
- Trames similaire à Excel PLX-DAQ: CLEARDATA / LABEL / RESETTIMER / DATA,TIMER,...
- Affichage direct: tableau à gauche, graphiques à droite
- Calculs: temps total, déformation max, température moyenne, vitesse de déformation (régression linéaire)
- Export: CSV, Excel
- Rapport: génération PDF directe (via ReportLab)
- Vidéo de théorie intégrée (fichier local)
- Ajout d'une protection à la fermeture pendant une acquisition (mot de passe)
- Disponible sous Linux et MacOS
- Ajout d'un logbook horodaté pour audit trail (actions utilisateur, erreurs, exportations de fichiers avec checksum)

------------------------------------------------------------------------------------------------------------

Auteurs: A. D. - D. F. - V. F-L. - T. M. (Armand Dreuslin, Dorian Fillion, Victor Fournier-Labé, Timoté Muller)
"""

from __future__ import annotations
from typing import Optional, TYPE_CHECKING, Any
from pathlib import Path
from dataclasses import dataclass
from datetime import datetime
from typing import List, Tuple

import os
import sys
import csv
import time
import uuid
import queue
import socket
import getpass
import hashlib
import platform
import traceback
import threading
import subprocess
import webbrowser

# --- GUI ---
try:
    import ttkbootstrap as tb
    from ttkbootstrap.constants import *
except Exception:
    tb = None

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog

# --- Serial ---
try:
    import serial
    from serial.tools import list_ports
except Exception:
    serial = None
    list_ports = None

if TYPE_CHECKING:
    from serial import Serial
else:
    Serial = Any

# --- Plot ---
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

APP_NAME = "FluageAutomation"
APP_VERSION = "1.07.001"
CLOSE_PASSWORD = APP_NAME + APP_VERSION


# ==================================================================================================
# RESSOURCES / COMPATIBILITÉ
# ==================================================================================================

def resource_path(*parts: str) -> str:
    """
    Retourne le chemin vers une ressource (assets, pdf, etc.)
    Compatible :
      - mode dev (script .py)
      - exe PyInstaller onedir : ressources souvent dans _internal/
    """
    if getattr(sys, "frozen", False):
        base = Path(sys.executable).resolve().parent
    else:
        base = Path(__file__).resolve().parent

    p = base.joinpath(*parts)
    if p.exists():
        return str(p)

    p2 = base.joinpath("_internal", *parts)
    return str(p2)


def get_os_name() -> str:
    return platform.system() or "Unknown"


def ui_font() -> str:
    system = get_os_name()
    if system == "Windows":
        return "Segoe UI"
    if system == "Darwin":
        return "SF Pro Text"
    return "DejaVu Sans"


def get_current_username() -> str:
    return (
        os.environ.get("USERNAME")
        or os.environ.get("USER")
        or getpass.getuser()
        or ""
    )


def maximize_window(win: tk.Tk | tk.Toplevel) -> None:
    system = get_os_name()
    try:
        if system == "Windows":
            win.state("zoomed")
        elif system == "Linux":
            win.attributes("-zoomed", True)
        else:
            win.geometry("1400x900")
    except Exception:
        try:
            win.geometry("1400x900")
        except Exception:
            pass


def open_with_default_app(path: str) -> None:
    system = get_os_name()
    if system == "Windows":
        os.startfile(path)  # type: ignore[attr-defined]
    elif system == "Darwin":
        subprocess.run(["open", path], check=False)
    else:
        subprocess.run(["xdg-open", path], check=False)


def open_iut_website():
    webbrowser.open("https://portail.univ-rouen.fr/")


def open_help():
    help_path = resource_path("assets", "aide.pdf")

    if not os.path.exists(help_path):
        messagebox.showerror(
            "Aide introuvable",
            f"Le fichier d'aide est introuvable :\n{help_path}\n\n"
            "Vérifie qu'il existe dans le dossier assets (ou qu'il est bien inclus dans l'exe)."
        )
        return

    open_with_default_app(help_path)


def open_video_theorie(self):
    video = Path(resource_path("assets", "Le fluage - Samuel Rey-Mermet.mp4"))
    if not video.exists():
        messagebox.showerror(
            "Vidéo introuvable",
            f"Le fichier vidéo est introuvable :\n{video}\n\n"
            "Vérifie qu'il existe dans le dossier."
        )
        return
    open_with_default_app(str(video))


ASSETS_DIR = resource_path("assets")
LOGO_IUT_PATH = resource_path("assets", "logoIutEvreux.jpg")
LOGO_APP_PATH = resource_path("assets", "app.png")
LOGO_GR_PATH = resource_path("assets", "group.png")

IUT_DARK = "#1D1D1B"
IUT_GREEN = "#A3C449"
IUT_GREEN_DARK = "#798F35"
BG_LIGHT = "#F6F7F9"


# ==================================================================================================
# LOGBOOK / AUDIT TRAIL
# ==================================================================================================

def get_machine_name() -> str:
    try:
        return socket.gethostname()
    except Exception:
        return "unknown-host"


def get_local_ip() -> str:
    """
    Essaie de récupérer l'IP locale de la machine sans dépendre d'un accès Internet réel.
    """
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        try:
            return socket.gethostbyname(socket.gethostname())
        except Exception:
            return "unknown-ip"


def sha256_file(path: str | Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


class Logbook:
    """
    Logbook texte horodaté pour audit trail simple.
    Ne doit jamais bloquer l'application si l'écriture échoue.
    """

    def __init__(self):
        self.start_dt = datetime.now()
        self.session_id = str(uuid.uuid4())
        self.username = get_current_username()
        self.os_name = get_os_name()
        self.machine_name = get_machine_name()
        self.machine_ip = get_local_ip()

        filename = self.start_dt.strftime("%d-%m-%Y_%H-%M-%S") + "_fluage-automation.txt"

        base_dir = Path(resource_path("."))
        self.log_dir = base_dir / "logs"
        self.log_dir.mkdir(parents=True, exist_ok=True)

        self.filepath = self.log_dir / filename
        self._write_header()

    def _safe_append(self, text: str) -> None:
        try:
            with open(self.filepath, "a", encoding="utf-8", newline="") as f:
                f.write(text)
                f.flush()
                os.fsync(f.fileno())
        except Exception:
            pass

    def _write_header(self) -> None:
        lines = [
            "===== FLUAGE AUTOMATION LOGBOOK =====",
            f"Lancement           : {self.start_dt.strftime('%d/%m/%Y %H:%M:%S')}",
            f"Application         : {APP_NAME}",
            f"Version             : {APP_VERSION}",
            f"Session ID          : {self.session_id}",
            f"Utilisateur         : {self.username}",
            f"Système             : {self.os_name}",
            f"Machine             : {self.machine_name}",
            f"IP machine          : {self.machine_ip}",
            "=" * 40,
            "",
        ]
        self._safe_append("\n".join(lines) + "\n")

    def write(self, action: str, details: str = "") -> None:
        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        line = f"[{timestamp}] | {action}"
        if details:
            line += f" | {details}"
        self._safe_append(line + "\n")

    def info(self, action: str, details: str = "") -> None:
        self.write(action, details)

    def warning(self, action: str, details: str = "") -> None:
        self.write(f"WARNING::{action}", details)

    def error(self, action: str, exc: Exception | str, details: str = "") -> None:
        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        text = str(exc)
        block = [
            f"[{timestamp}] | ERROR::{action}" + (f" | {details}" if details else ""),
            f"Type       : {type(exc).__name__ if isinstance(exc, Exception) else 'Error'}",
            f"Message    : {text}",
        ]
        if isinstance(exc, Exception):
            block.append("Traceback   :")
            block.append(traceback.format_exc().rstrip())
        block.append("-" * 100)
        self._safe_append("\n".join(block) + "\n")

    def file_export(self, action: str, filepath: str | Path) -> None:
        try:
            filepath = Path(filepath)
            checksum = sha256_file(filepath)
            size = filepath.stat().st_size
            self.write(
                action,
                f"File={filepath} | Size={size} bytes | SHA256={checksum}"
            )
        except Exception as e:
            self.error(f"{action}_CHECKSUM", e, f"File={filepath}")


# ==================================================================================================
# DONNÉES
# ==================================================================================================

@dataclass
class MeasureRow:
    t_s: float
    distance_mm: float
    elong_mm: float
    strain_pct: float
    temp_c: float


# ==================================================================================================
# WORKER SÉRIE
# ==================================================================================================

class SerialWorker:
    """Thread de lecture série -> pousse les lignes dans une Queue."""

    def __init__(self, port: str, baud: int = 115200, timeout: float = 0.5):
        if serial is None:
            raise RuntimeError("pyserial n'est pas installé")
        self.port = port
        self.baud = baud
        self.timeout = timeout
        self._ser: Optional[Serial] = None
        self._thread: Optional[threading.Thread] = None
        self._stop = threading.Event()
        self.lines: "queue.Queue[str]" = queue.Queue()

    def start(self) -> None:
        self._stop.clear()
        self._ser = serial.Serial(self.port, self.baud, timeout=self.timeout)
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()

    def stop(self) -> None:
        self._stop.set()
        if self._ser and self._ser.is_open:
            try:
                self._ser.close()
            except Exception:
                pass
        self._ser = None

    def send_line(self, text: str) -> None:
        if self._ser and self._ser.is_open:
            self._ser.write((text.strip() + "\n").encode("utf-8"))
            self._ser.flush()

    def reset_input_buffer(self) -> None:
        if self._ser and self._ser.is_open:
            try:
                self._ser.reset_input_buffer()
            except Exception:
                pass

    def _run(self) -> None:
        assert self._ser is not None
        while not self._stop.is_set():
            try:
                raw = self._ser.readline()
                if not raw:
                    continue
                line = raw.decode(errors="ignore").strip()
                if line:
                    self.lines.put(line)
            except Exception:
                time.sleep(0.1)


# ==================================================================================================
# PLOTS
# ==================================================================================================

class LivePlot:
    """2 graphiques live: déformation(t) et température(t)."""

    def __init__(self, master: tk.Widget):
        self.container = ttk.Frame(master)

        toolbar = ttk.Frame(self.container)
        toolbar.pack(side=tk.TOP, fill=tk.X, padx=5, pady=(0, 5))

        ttk.Label(toolbar, text="Agrandir :", font=(ui_font(), 9)).pack(side=tk.LEFT, padx=(0, 10))

        self.btn_zoom_strain = ttk.Button(
            toolbar,
            text="🔍 Déformation",
            command=lambda: self.toggle_zoom("strain"),
            width=18
        )
        self.btn_zoom_strain.pack(side=tk.LEFT, padx=2)

        self.btn_zoom_temp = ttk.Button(
            toolbar,
            text="🔍 Température",
            command=lambda: self.toggle_zoom("temp"),
            width=18
        )
        self.btn_zoom_temp.pack(side=tk.LEFT, padx=2)

        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)

        ttk.Label(toolbar, text="Curseurs :", font=(ui_font(), 9)).pack(side=tk.LEFT, padx=(0, 10))

        self.btn_toggle_cursors = ttk.Button(
            toolbar,
            text="📍 Placer",
            command=self.toggle_cursors,
            width=12
        )
        self.btn_toggle_cursors.pack(side=tk.LEFT, padx=2)

        self.lbl_cursor_info = ttk.Label(toolbar, text="", font=(ui_font(), 9), foreground="#6B7280")
        self.lbl_cursor_info.pack(side=tk.LEFT, padx=10)

        self.fig = Figure(dpi=100, constrained_layout=True)
        self.ax1 = self.fig.add_subplot(211)
        self.ax2 = self.fig.add_subplot(212)

        self.ax1.set_title("Déformation en fonction du temps")
        self.ax1.set_xlabel("Temps (s)")
        self.ax1.set_ylabel("Déformation (%)")
        self.ax1.grid(True)

        self.ax2.set_title("Température en fonction du temps")
        self.ax2.set_xlabel("Temps (s)")
        self.ax2.set_ylabel("Température (°C)")
        self.ax2.grid(True)

        (self.line_strain,) = self.ax1.plot([], [])
        (self.line_temp,) = self.ax2.plot([], [])
        (self.line_trend,) = self.ax1.plot([], [], linestyle="--", color="red", linewidth=1.5)
        self.eq_text = self.ax1.text(
            0.98, 0.02, "", transform=self.ax1.transAxes,
            ha="right", va="bottom", fontsize=12, color="red"
        )

        self._trend: dict | None = None

        self.canvas = FigureCanvasTkAgg(self.fig, master=self.container)
        self.widget = self.canvas.get_tk_widget()
        self.widget.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self._zoom: dict[str, dict] = {}

        self._cursors_active = False
        self._cursor_lines = []
        self._cursor_positions = []
        self._dragging_cursor = None
        self._cursor_callback = None

        self.canvas.mpl_connect("button_press_event", self._on_cursor_press)
        self.canvas.mpl_connect("button_release_event", self._on_cursor_release)
        self.canvas.mpl_connect("motion_notify_event", self._on_cursor_motion)

        self._x: List[float] = []
        self._strain: List[float] = []
        self._temp: List[float] = []

        self._time_factor: float = 1.0
        self._time_label: str = "s"

    def clear(self) -> None:
        self._x.clear()
        self._strain.clear()
        self._temp.clear()
        self._trend = None
        self.line_trend.set_data([], [])
        self.eq_text.set_text("")
        self._redraw(full=True)
        self._update_zoom("strain", full=True)
        self._update_zoom("temp", full=True)

    def append(self, t: float, strain_pct: float, temp_c: float) -> None:
        self._x.append(t)
        self._strain.append(strain_pct)
        self._temp.append(temp_c)
        self._redraw(full=False)
        self._update_zoom("strain", full=False)
        self._update_zoom("temp", full=False)

    def get_arrays(self) -> Tuple[List[float], List[float], List[float]]:
        return self._x, self._strain, self._temp

    def set_time_unit(self, unit: str) -> None:
        unit = (unit or "s").lower()
        if unit in ("s", "sec", "seconde", "secondes"):
            self._time_factor = 1.0
            self._time_label = "s"
        elif unit in ("min", "m", "minute", "minutes"):
            self._time_factor = 60.0
            self._time_label = "min"
        elif unit in ("h", "hr", "heure", "heures"):
            self._time_factor = 3600.0
            self._time_label = "h"
        else:
            self._time_factor = 1.0
            self._time_label = "s"

        self._update_zoom("strain", full=True)
        self._update_zoom("temp", full=True)

        self.ax1.set_xlabel(f"Temps ({self._time_label})")
        self.ax2.set_xlabel(f"Temps ({self._time_label})")
        self._redraw(full=True)

    def get_time_display_params(self) -> Tuple[float, str]:
        return self._time_factor, self._time_label

    def set_trendline(self, a: float, b: float, t0_s: float, t1_s: float, r2: float | None = None) -> None:
        self._trend = {"a": a, "b": b, "t0": t0_s, "t1": t1_s, "r2": r2}
        self._redraw(full=True)

    def clear_trendline(self) -> None:
        self._trend = None
        self.line_trend.set_data([], [])
        self.eq_text.set_text("")
        self.canvas.draw_idle()

    def set_cursor_callback(self, callback) -> None:
        self._cursor_callback = callback

    def toggle_cursors(self) -> None:
        if self._cursors_active:
            self._remove_cursors()
        else:
            self._create_cursors()

    def _create_cursors(self) -> None:
        if not self._x:
            return

        t_min, t_max = min(self._x), max(self._x)
        t0 = t_min + 0.25 * (t_max - t_min)
        t1 = t_min + 0.75 * (t_max - t_min)

        self._cursor_positions = [t0, t1]

        for t_s in self._cursor_positions:
            t_display = t_s / self._time_factor
            line = self.ax1.axvline(t_display, color="orange", linewidth=2, alpha=0.7, linestyle="--", picker=5)
            self._cursor_lines.append(line)

        self._cursors_active = True
        self.btn_toggle_cursors.config(text="❌ Supprimer")
        self._update_cursor_label()
        self.canvas.draw_idle()

        if self._cursor_callback:
            self._cursor_callback(self._cursor_positions[0], self._cursor_positions[1])

    def _remove_cursors(self) -> None:
        for line in self._cursor_lines:
            line.remove()
        self._cursor_lines.clear()
        self._cursor_positions.clear()
        self._cursors_active = False
        self._dragging_cursor = None
        self.btn_toggle_cursors.config(text="📍 Placer")
        self.lbl_cursor_info.config(text="")
        self.canvas.draw_idle()

    def _update_cursor_label(self) -> None:
        if not self._cursor_positions:
            self.lbl_cursor_info.config(text="")
            return

        t0, t1 = self._cursor_positions
        t0_disp = t0 / self._time_factor
        t1_disp = t1 / self._time_factor
        text = f"t₀ = {t0_disp:.2f} {self._time_label}  |  t₁ = {t1_disp:.2f} {self._time_label}"
        self.lbl_cursor_info.config(text=text)

    def _on_cursor_press(self, event) -> None:
        if not self._cursors_active or event.inaxes != self.ax1:
            return
        for i, line in enumerate(self._cursor_lines):
            if line.contains(event)[0]:
                self._dragging_cursor = i
                break

    def _on_cursor_release(self, event) -> None:
        if self._dragging_cursor is not None:
            self._dragging_cursor = None
            if self._cursor_callback:
                self._cursor_callback(self._cursor_positions[0], self._cursor_positions[1])

    def _on_cursor_motion(self, event) -> None:
        if self._dragging_cursor is None or event.inaxes != self.ax1 or event.xdata is None:
            return

        t_new_s = event.xdata * self._time_factor

        if self._x:
            t_new_s = max(min(self._x), min(t_new_s, max(self._x)))

        idx = self._dragging_cursor
        if idx == 0:
            t_new_s = min(t_new_s, self._cursor_positions[1] - 0.1)
        else:
            t_new_s = max(t_new_s, self._cursor_positions[0] + 0.1)

        self._cursor_positions[idx] = t_new_s
        t_display = t_new_s / self._time_factor
        self._cursor_lines[idx].set_xdata([t_display, t_display])

        self._update_cursor_label()
        self.canvas.draw_idle()

    def _redraw(self, full: bool) -> None:
        x_view = [v / self._time_factor for v in self._x]
        self.line_strain.set_data(x_view, self._strain)
        self.line_temp.set_data(x_view, self._temp)

        if self._trend is not None:
            t0 = float(self._trend["t0"])
            t1 = float(self._trend["t1"])
            a = float(self._trend["a"])
            b = float(self._trend["b"])
            r2 = self._trend.get("r2", None)
            x_seg_s = [t0, t1]
            y_seg = [a * t0 + b, a * t1 + b]
            x_seg = [v / self._time_factor for v in x_seg_s]
            self.line_trend.set_data(x_seg, y_seg)

            a_disp = a * self._time_factor
            eq = f"y = {a_disp:.6g} x + {b:.6g}"
            if r2 is not None:
                eq += f"\nR² = {float(r2):.4f}"
            self.eq_text.set_text(eq)
        else:
            self.line_trend.set_data([], [])
            self.eq_text.set_text("")

        if full or (len(self._x) % 5 == 0):
            for ax, ys in [(self.ax1, self._strain), (self.ax2, self._temp)]:
                ax.relim()
                ax.autoscale_view()

        self.canvas.draw_idle()

    def toggle_zoom(self, which: str) -> None:
        which = (which or "").lower().strip()
        if which not in ("strain", "temp"):
            return

        z = self._zoom.get(which)
        if z and z.get("win") and z["win"].winfo_exists():
            try:
                z["win"].destroy()
            except Exception:
                pass
            self._zoom.pop(which, None)
            return

        self._open_zoom(which)

    def _open_zoom(self, which: str) -> None:
        win = tk.Toplevel(self.widget)
        win.title("Zoom — Déformation" if which == "strain" else "Zoom — Température")
        win.minsize(900, 600)

        fig = Figure(dpi=110, constrained_layout=True)
        ax = fig.add_subplot(111)

        if which == "strain":
            ax.set_title("Déformation en fonction du temps")
            ax.set_xlabel(f"Temps ({self._time_label})")
            ax.set_ylabel("Déformation (%)")
            ax.grid(True)
            (line_main,) = ax.plot([], [], linewidth=2.0)
            (line_trend,) = ax.plot([], [], linestyle="--", color="red", linewidth=2.0)
            eq_text = ax.text(
                0.98, 0.02, "", transform=ax.transAxes,
                ha="right", va="bottom", fontsize=12, color="red"
            )
        else:
            ax.set_title("Température en fonction du temps")
            ax.set_xlabel(f"Temps ({self._time_label})")
            ax.set_ylabel("Température (°C)")
            ax.grid(True)
            (line_main,) = ax.plot([], [], linewidth=2.0)
            line_trend = None
            eq_text = None

        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.get_tk_widget().pack(fill="both", expand=True)

        self._zoom[which] = {
            "win": win,
            "fig": fig,
            "ax": ax,
            "canvas": canvas,
            "line_main": line_main,
            "line_trend": line_trend,
            "eq_text": eq_text,
        }

        def _on_close():
            try:
                win.destroy()
            finally:
                self._zoom.pop(which, None)

        win.protocol("WM_DELETE_WINDOW", _on_close)
        self._update_zoom(which, full=True)

    def _update_zoom(self, which: str, full: bool = False) -> None:
        z = self._zoom.get(which)
        if not z or not z.get("win") or not z["win"].winfo_exists():
            return

        x_view = [v / self._time_factor for v in self._x]

        if which == "strain":
            z["line_main"].set_data(x_view, self._strain)

            if self._trend is not None:
                t0 = float(self._trend["t0"])
                t1 = float(self._trend["t1"])
                a = float(self._trend["a"])
                b = float(self._trend["b"])
                r2 = self._trend.get("r2", None)

                x_seg = [t0 / self._time_factor, t1 / self._time_factor]
                y_seg = [a * t0 + b, a * t1 + b]
                z["line_trend"].set_data(x_seg, y_seg)

                a_disp = a * self._time_factor
                eq = f"y = {a_disp:.6g} x + {b:.6g}"
                if r2 is not None:
                    eq += f"\nR² = {float(r2):.4f}"
                if z["eq_text"] is not None:
                    z["eq_text"].set_text(eq)
            else:
                if z["line_trend"] is not None:
                    z["line_trend"].set_data([], [])
                if z["eq_text"] is not None:
                    z["eq_text"].set_text("")

            z["ax"].set_xlabel(f"Temps ({self._time_label})")

        else:
            z["line_main"].set_data(x_view, self._temp)
            z["ax"].set_xlabel(f"Temps ({self._time_label})")

        if full or (len(self._x) % 5 == 0):
            z["ax"].relim()
            z["ax"].autoscale_view()

        z["canvas"].draw_idle()


# ==================================================================================================
# APPLICATION PRINCIPALE
# ==================================================================================================

class FluageApp:
    def __init__(self):
        if tb is None:
            raise RuntimeError("ttkbootstrap n'est pas installé. Fais: pip install ttkbootstrap")

        self.log = Logbook()
        self.log.info("APPLICATION_START", "Initialisation de l'application")

        self.root = tb.Window(themename="flatly")
        self.root.title(f"{APP_NAME} — v{APP_VERSION}")

        self._app_icon = None
        try:
            ico_path = resource_path("assets", "app.ico")
            png_path = resource_path("assets", "app.png")

            if get_os_name() == "Windows" and os.path.exists(ico_path):
                self.root.iconbitmap(ico_path)

            if os.path.exists(png_path):
                self._app_icon = tk.PhotoImage(file=png_path)
                self.root.iconphoto(True, self._app_icon)
        except Exception as e:
            self.log.error("ICON_LOAD", e)

        maximize_window(self.root)
        self.root.minsize(1200, 800)

        style = tb.Style()
        try:
            style.colors.primary = IUT_GREEN
            style.colors.dark = IUT_DARK
        except Exception as e:
            self.log.error("STYLE_INIT", e)

        self.worker: Optional[SerialWorker] = None
        self.rows: List[MeasureRow] = []

        self.time_unit = tk.StringVar(value="s")
        self._last_slope_pct_per_s: float | None = None

        try:
            self._build_layout()
            self._refresh_ports()
            self._ui_tick()
            self.cbo_port.bind("<<ComboboxSelected>>", self._update_port_label)
            self.root.protocol("WM_DELETE_WINDOW", self._on_app_close)
        except Exception as e:
            self.log.error("APP_INIT", e)
            raise

    def _on_app_close(self) -> None:
        if self.worker is not None:
            pwd = simpledialog.askstring(
                "Fermeture protégée",
                "Une acquisition est en cours.\n"
                "Pour fermer le logiciel, entrez le mot de passe :",
                show="*",
                parent=self.root
            )

            if pwd is None:
                self.status.set("Fermeture annulée")
                self.log.warning("APPLICATION_CLOSE_CANCELLED", "Annulation utilisateur pendant acquisition")
                return

            if pwd != CLOSE_PASSWORD:
                messagebox.showerror(
                    "Mot de passe incorrect",
                    "Mot de passe incorrect.\nLa fermeture du logiciel a été bloquée."
                )
                self.status.set("Tentative de fermeture bloquée")
                self.log.warning("APPLICATION_CLOSE_BLOCKED", "Mot de passe incorrect")
                return

        if self.worker:
            try:
                self.worker.send_line("STOP")
                time.sleep(0.2)
            except Exception as e:
                self.log.error("SERIAL_STOP_SEND", e)

            try:
                self.worker.stop()
            except Exception as e:
                self.log.error("SERIAL_STOP", e)

            self.worker = None

        self.log.info("APPLICATION_CLOSE", "Fermeture normale")
        self.root.destroy()

    def _build_layout(self) -> None:
        header = tb.Frame(self.root, padding=(14, 10))
        header.pack(side=tk.TOP, fill=tk.X)

        btn_video = tk.Button(
            header,
            text="🎬 Vidéo théorie",
            command=self._open_video_logged,
            borderwidth=0,
        )
        btn_video.pack(side=tk.LEFT, padx=(0, 12))

        self._logo_img = None
        if os.path.exists(LOGO_IUT_PATH):
            try:
                from PIL import Image, ImageTk

                imGroup = Image.open(LOGO_GR_PATH)
                imGroup = imGroup.resize((80, int(80 * imGroup.height / imGroup.width)))
                self._logo_group_img = ImageTk.PhotoImage(imGroup)

                btn_help = tk.Button(
                    header,
                    image=self._logo_group_img,
                    command=self._open_help_logged,
                    borderwidth=0,
                    highlightthickness=0,
                    relief="flat",
                    cursor="hand2"
                )
                btn_help.pack(side=tk.RIGHT, padx=(30, 12))

                im = Image.open(LOGO_IUT_PATH)
                im = im.resize((250, int(250 * im.height / im.width)))
                self._logo_img = ImageTk.PhotoImage(im)

                btn_iut = tk.Button(
                    header,
                    image=self._logo_img,
                    command=self._open_iut_logged,
                    borderwidth=0,
                    highlightthickness=0,
                    relief="flat",
                    cursor="hand2"
                )
                btn_iut.pack(side=tk.RIGHT)

                imApp = Image.open(LOGO_APP_PATH)
                imApp = imApp.resize((80, int(80 * imApp.height / imApp.width)))
                self._logo_app_img = ImageTk.PhotoImage(imApp)
                tb.Label(header, image=self._logo_app_img).pack(side=tk.LEFT, padx=(0, 12))
            except Exception as e:
                self.log.error("LOGO_LOAD", e)

        title_box = tb.Frame(header)
        title_box.pack(side=tk.LEFT, padx=14)
        tb.Label(title_box, text=APP_NAME, font=(ui_font(), 18, "bold"), foreground=IUT_DARK).pack(anchor="w")
        tb.Label(
            title_box,
            text="Acquisition série pour essais de fluage",
            font=(ui_font(), 10),
            foreground="#4b5563",
        ).pack(anchor="w")

        main = tb.Panedwindow(self.root, orient=tk.HORIZONTAL)
        main.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))

        left = tb.Frame(main, padding=10)
        right = tb.Frame(main, padding=10)
        main.add(left, weight=1)
        main.add(right, weight=2)

        ctrl = tb.Labelframe(left, text="Connexion & Actions", padding=10)
        ctrl.pack(fill=tk.X)

        row1 = tb.Frame(ctrl)
        row1.pack(fill=tk.X)

        tb.Label(row1, text="♾️ Port série").pack(side=tk.LEFT)
        self.cbo_port = tb.Combobox(row1, width=22, state="readonly")
        self.cbo_port.pack(side=tk.LEFT, padx=8)

        tb.Label(row1, text="Nom :").pack(side=tk.LEFT, padx=(4, 0))
        self.lbl_port_name = tb.Label(row1, text="", bootstyle="secondary")
        self.lbl_port_name.pack(side=tk.LEFT, padx=(4, 12))

        tb.Label(row1, text="Bauds :").pack(side=tk.LEFT, padx=(4, 0))
        self.fixed_baud = 115200
        self.lbl_baud = tb.Label(row1, text=str(self.fixed_baud), bootstyle="secondary")
        self.lbl_baud.pack(side=tk.LEFT, padx=8)

        self.btn_refresh = tb.Button(row1, text="⟳ Rafraîchir", bootstyle="secondary", command=self._refresh_ports)
        self.btn_refresh.pack(side=tk.RIGHT)

        row2 = tb.Frame(ctrl)
        row2.pack(fill=tk.X, pady=(10, 0))

        self.btn_connect = tb.Button(row2, text="🔌 Connexion", bootstyle="success", command=self._connect)
        self.btn_connect.pack(side=tk.LEFT)
        self.btn_disconnect = tb.Button(row2, text="⏻ Déconnexion", bootstyle="danger", command=self._disconnect, state=tk.DISABLED)
        self.btn_disconnect.pack(side=tk.LEFT, padx=8)

        self.btn_clear = tb.Button(row2, text="🧹 Vider", bootstyle="warning", command=self._clear)
        self.btn_clear.pack(side=tk.LEFT, padx=8)

        row3 = tb.Frame(ctrl)
        row3.pack(fill=tk.X, pady=(10, 0))

        self.btn_export_csv = tb.Button(row3, text="📥 Exporter CSV", bootstyle="primary", command=self._export_csv)
        self.btn_export_csv.pack(side=tk.LEFT)
        self.btn_export_xlsx = tb.Button(row3, text="📥 Exporter Excel", bootstyle="primary", command=self._export_xlsx)
        self.btn_export_xlsx.pack(side=tk.LEFT, padx=8)
        self.btn_report = tb.Button(row3, text="📄 Générer rapport", bootstyle="primary", command=self._generate_report)
        self.btn_report.pack(side=tk.RIGHT)

        summary = tb.Labelframe(left, text="Résumé", padding=10)
        summary.pack(fill=tk.X, pady=(10, 0))

        unit_row = tb.Frame(summary)
        unit_row.pack(fill=tk.X, pady=(0, 6))
        tb.Label(unit_row, text="Axe temps :", foreground="#4b5563").pack(side=tk.LEFT)

        for txt_label, val in [("Secondes", "s"), ("Minutes", "min"), ("Heures", "h")]:
            tb.Radiobutton(
                unit_row,
                text=txt_label,
                value=val,
                variable=self.time_unit,
                command=self._on_time_unit_change,
            ).pack(side=tk.LEFT, padx=(8, 0))

        self.var_time = tk.StringVar(value="—")
        self.var_strain = tk.StringVar(value="—")
        self.var_temp = tk.StringVar(value="—")
        self.var_rate = tk.StringVar(value="—")

        grid = tb.Frame(summary)
        grid.pack(fill=tk.X)

        def metric(r, label, var):
            tb.Label(grid, text=label, font=(ui_font(), 10, "bold"), foreground=IUT_DARK).grid(row=r, column=0, sticky="w")
            tb.Label(grid, textvariable=var, font=(ui_font(), 10), foreground="#111827").grid(row=r, column=1, sticky="w", padx=(22, 0))

        metric(0, "Temps total", self.var_time)
        metric(1, "Déformation max", self.var_strain)
        metric(2, "Température moyenne", self.var_temp)
        metric(3, "Vitesse de déformation", self.var_rate)

        grid.columnconfigure(0, weight=0)
        grid.columnconfigure(1, weight=0)
        grid.columnconfigure(2, weight=1)
        grid.columnconfigure(1, weight=1)

        rate_box = tb.Labelframe(left, text="Vitesse de déformation (régression linéaire)", padding=10)
        rate_box.pack(fill=tk.X, pady=(10, 0))

        rate_row = tb.Frame(rate_box)
        rate_row.pack(fill=tk.X)

        self.var_auto_rate = tk.BooleanVar(value=True)
        tb.Checkbutton(
            rate_row,
            text="Auto (zone la plus représentative)",
            variable=self.var_auto_rate,
            command=self._on_toggle_auto_rate,
        ).pack(side=tk.LEFT, padx=(0, 12))

        tb.Label(rate_row, text="t début (s)").pack(side=tk.LEFT)
        self.ent_t0 = tb.Entry(rate_row, width=10)
        self.ent_t0.insert(0, "0")
        self.ent_t0.pack(side=tk.LEFT, padx=8)

        tb.Label(rate_row, text="t fin (s)").pack(side=tk.LEFT)
        self.ent_t1 = tb.Entry(rate_row, width=10)
        self.ent_t1.insert(0, "60")
        self.ent_t1.pack(side=tk.LEFT, padx=8)

        self.btn_rate = tb.Button(rate_row, text="📈 Calculer", bootstyle="success", command=self._compute_strain_rate)
        self.btn_rate.pack(side=tk.RIGHT)

        table_box = tb.Labelframe(left, text="Mesures (live)", padding=8)
        table_box.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        cols = ("t", "dist", "elong", "strain", "temp")
        self.tree = ttk.Treeview(table_box, columns=cols, show="headings", height=15)
        self.tree.heading("t", text="Temps (s)")
        self.tree.heading("dist", text="Distance (mm)")
        self.tree.heading("elong", text="Allongement (mm)")
        self.tree.heading("strain", text="Déformation (%)")
        self.tree.heading("temp", text="Température (°C)")

        for c, w in [("t", 90), ("dist", 110), ("elong", 130), ("strain", 130), ("temp", 130)]:
            self.tree.column(c, width=w, anchor="center")

        yscroll = ttk.Scrollbar(table_box, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)

        plot_box = tb.Labelframe(right, text="Graphiques (auto)", padding=10)
        plot_box.pack(fill=tk.BOTH, expand=True)

        self.plot = LivePlot(plot_box)
        self.plot.container.pack(fill=tk.BOTH, expand=True)
        self.plot.set_cursor_callback(self._on_cursor_moved)

        self.status = tk.StringVar(value="Prêt")
        sb = tb.Frame(self.root, padding=(12, 6))
        sb.pack(side=tk.BOTTOM, fill=tk.X)
        tb.Label(sb, textvariable=self.status, foreground="#374151").pack(side=tk.LEFT)

    def _open_help_logged(self):
        try:
            open_help()
            self.log.info("OPEN_HELP")
        except Exception as e:
            self.log.error("OPEN_HELP", e)

    def _open_video_logged(self):
        try:
            open_video_theorie(self)
            self.log.info("OPEN_VIDEO_THEORIE")
        except Exception as e:
            self.log.error("OPEN_VIDEO_THEORIE", e)

    def _open_iut_logged(self):
        try:
            open_iut_website()
            self.log.info("OPEN_IUT_WEBSITE")
        except Exception as e:
            self.log.error("OPEN_IUT_WEBSITE", e)

    # ------------- Ports / Connect -------------
    def _refresh_ports(self) -> None:
        if serial is None or list_ports is None:
            self.cbo_port["values"] = []
            self.lbl_port_name.config(text="")
            self.status.set("pyserial manquant")
            self.log.warning("REFRESH_PORTS", "pyserial manquant")
            return

        priority_ports = []
        other_ports = []
        self._ports_info = {}

        try:
            for p in list_ports.comports():
                device = (p.device or "").strip()
                description = (p.description or "Périphérique inconnu").strip()

                if not device:
                    continue

                self._ports_info[device] = description
                dev_lower = device.lower()

                if any(k in dev_lower for k in ("ttyusb", "ttyacm", "usbmodem", "usbserial", "com")):
                    priority_ports.append(device)
                else:
                    other_ports.append(device)

            ports = list(dict.fromkeys(priority_ports + other_ports))
            self.cbo_port["values"] = ports

            current = self.cbo_port.get().strip()

            if ports:
                if current not in ports:
                    self.cbo_port.set(ports[0])
                self._update_port_label()
                self.status.set(f"{len(ports)} port(s) série détecté(s)")
                self.log.info("REFRESH_PORTS", f"Ports détectés={ports}")
            else:
                self.cbo_port.set("")
                self.lbl_port_name.config(text="")
                self.status.set("Aucun port série détecté")
                self.log.info("REFRESH_PORTS", "Aucun port détecté")
        except Exception as e:
            self.log.error("REFRESH_PORTS", e)

    def _update_port_label(self, event=None):
        port = self.cbo_port.get()
        name = self._ports_info.get(port, "")
        self.lbl_port_name.config(text=name)

    def _connect(self) -> None:
        selected = self.cbo_port.get().strip()
        if not selected:
            messagebox.showerror("Connexion", "Choisis un port série.")
            self.log.warning("CONNECT", "Aucun port sélectionné")
            return

        port = getattr(self, "_ports_map", {}).get(selected, selected)

        if "—" in port:
            port = port.split("—", 1)[0].strip()

        baud = getattr(self, "fixed_baud", 115200)
        try:
            self.worker = SerialWorker(port, baud=baud)
            self.worker.start()
            time.sleep(2.0)
            self.worker.reset_input_buffer()
            self.worker.send_line("START")
        except Exception as e:
            extra = ""
            if get_os_name() == "Linux":
                extra = "\n\nSous Linux, vérifie aussi les droits d'accès au port série (/dev/ttyUSB0, /dev/ttyACM0, etc.)."
            messagebox.showerror("Connexion", f"Impossible d'ouvrir {port}:\n{e}{extra}")
            self.log.error("CONNECT", e, f"Port={port} | Baud={baud}")
            self.worker = None
            return

        self.btn_connect.configure(state=tk.DISABLED)
        self.btn_disconnect.configure(state=tk.NORMAL)
        self.status.set(f"Connecté à {port} @ {baud}")
        self._set_action_buttons_state(connected=True)
        self.log.info("CONNECT", f"Port={port} | Baud={baud}")

    def _disconnect(self) -> None:
        if self.worker:
            try:
                self.worker.send_line("STOP")
                time.sleep(0.2)
            except Exception as e:
                self.log.error("DISCONNECT_STOP_SEND", e)

            try:
                self.worker.stop()
            except Exception as e:
                self.log.error("DISCONNECT_STOP", e)

        self.worker = None

        self.btn_connect.configure(state=tk.NORMAL)
        self.btn_disconnect.configure(state=tk.DISABLED)
        self._set_action_buttons_state(connected=False)
        self.status.set("Déconnecté")
        self.log.info("DISCONNECT")

    def _set_action_buttons_state(self, *, connected: bool) -> None:
        state = tk.DISABLED if connected else tk.NORMAL
        for btn in (self.btn_clear, self.btn_export_csv, self.btn_export_xlsx, self.btn_report):
            try:
                btn.configure(state=state)
            except Exception as e:
                self.log.error("SET_BUTTON_STATE", e)

        port_state = tk.DISABLED if connected else "readonly"
        try:
            self.cbo_port.configure(state=port_state)
        except Exception as e:
            self.log.error("SET_PORT_STATE", e)

        try:
            self.btn_refresh.configure(state=state)
        except Exception as e:
            self.log.error("SET_REFRESH_STATE", e)

    # ------------- Data handling -------------
    def _clear_data(self) -> None:
        self.rows.clear()
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.plot.clear()

        if self.plot._cursors_active:
            self.plot._remove_cursors()

        self._update_summary()
        self.status.set("Données vidées")
        self.log.info("CLEAR_DATA")

    def _clear(self) -> None:
        if self.worker is not None:
            messagebox.showinfo("Action indisponible", "Déconnectez-vous avant de vider les données.")
            self.log.warning("CLEAR_REQUEST_BLOCKED", "Tentative pendant acquisition")
            return

        if messagebox.askyesno("Confirmer", "Êtes-vous sûr de vouloir vider toutes les données ?", parent=self.root):
            if messagebox.askyesno("Confirmation finale", "Cette action est irréversible. Confirmez-vous ?", parent=self.root):
                self._clear_data()
                messagebox.showinfo("Données vidées", "Toutes les données ont été vidées.")
            else:
                self.log.info("CLEAR_CANCELLED", "Confirmation finale refusée")
        else:
            self.log.info("CLEAR_CANCELLED", "Première confirmation refusée")

    def _ui_tick(self) -> None:
        if self.worker:
            try:
                for _ in range(50):
                    line = self.worker.lines.get_nowait()
                    self._handle_line(line)
            except queue.Empty:
                pass
            except Exception as e:
                self.log.error("UI_TICK", e)

        self.root.after(100, self._ui_tick)

    def _handle_line(self, line: str) -> None:
        if line.startswith("CLEARDATA"):
            self._clear_data()
            return
        if line.startswith("LABEL") or line.startswith("RESETTIMER"):
            return

        if not line.startswith("DATA"):
            return

        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 6:
            self.log.warning("SERIAL_LINE_INVALID", f"Ligne ignorée={line}")
            return

        try:
            if parts[1].upper() == "TIMER":
                t = 0.0 if not self.rows else self.rows[-1].t_s + 0.5
                dc = float(parts[2])
                dL = float(parts[3])
                eps = float(parts[4])
                temp = float(parts[5])
            else:
                t = float(parts[1])
                dc = float(parts[2])
                dL = float(parts[3])
                eps = float(parts[4])
                temp = float(parts[5])
        except Exception as e:
            self.log.error("SERIAL_PARSE", e, f"Ligne={line}")
            return

        row = MeasureRow(t, dc, dL, eps, temp)
        self.rows.append(row)

        self.tree.insert("", tk.END, values=(
            f"{t:.1f}",
            f"{dc:.3f}",
            f"{dL:.4f}",
            f"{eps:.4f}",
            f"{temp:.1f}",
        ))

        if len(self.tree.get_children()) > 3000:
            children = self.tree.get_children()
            excess = len(children) - 3000
            for iid in children[:min(excess, 100)]:
                self.tree.delete(iid)

        self.plot.append(t, eps, temp)

        if len(self.rows) % 5 == 0:
            self._update_summary()

    def _on_time_unit_change(self) -> None:
        unit = self.time_unit.get()
        self.plot.set_time_unit(unit)
        self._update_summary()
        self.log.info("TIME_UNIT_CHANGE", f"Unit={unit}")
        if self._last_slope_pct_per_s is not None:
            self._set_rate_display(self._last_slope_pct_per_s)

    def _on_cursor_moved(self, t0_s: float, t1_s: float) -> None:
        self.var_auto_rate.set(False)

        self.ent_t0.delete(0, tk.END)
        self.ent_t0.insert(0, f"{t0_s:.2f}")

        self.ent_t1.delete(0, tk.END)
        self.ent_t1.insert(0, f"{t1_s:.2f}")

        self.btn_rate.config(state="normal")
        self.log.info("CURSOR_MOVED", f"t0={t0_s:.2f} | t1={t1_s:.2f}")

    def _time_display(self, seconds: float) -> tuple[float, str]:
        unit = self.time_unit.get()
        if unit == "min":
            return seconds / 60.0, "min"
        if unit == "h":
            return seconds / 3600.0, "h"
        return seconds, "s"

    def _set_rate_display(self, slope_pct_per_s: float) -> None:
        unit = self.time_unit.get()
        if unit == "min":
            val = slope_pct_per_s * 60.0
            self.var_rate.set(f"{val:.6f} %/min")
        elif unit == "h":
            val = slope_pct_per_s * 3600.0
            self.var_rate.set(f"{val:.6f} %/h")
        else:
            self.var_rate.set(f"{slope_pct_per_s:.6f} %/s")

    def _update_summary(self) -> None:
        if not self.rows:
            self.var_time.set("—")
            self.var_strain.set("—")
            self.var_temp.set("—")
            self.var_rate.set("—")
            return

        t_tot = self.rows[-1].t_s
        eps_max = max(r.strain_pct for r in self.rows)
        temp_mean = sum(r.temp_c for r in self.rows) / len(self.rows)

        t_val, t_unit = self._time_display(t_tot)
        self.var_time.set(f"{t_val:.2f} {t_unit}")
        self.var_strain.set(f"{eps_max:.4f} %")
        self.var_temp.set(f"{temp_mean:.2f} °C")

    # ------------- Analytics -------------
    def _compute_strain_rate(self) -> None:
        if len(self.rows) < 3:
            messagebox.showinfo("Vitesse", "Pas assez de points.")
            self.log.warning("COMPUTE_RATE", "Pas assez de points")
            return

        if getattr(self, "var_auto_rate", None) is not None and self.var_auto_rate.get():
            best = self._auto_select_linear_window()
            if best is None:
                messagebox.showinfo("Vitesse", "Impossible de trouver une zone linéaire représentative.")
                self.log.warning("COMPUTE_RATE", "Aucune zone linéaire représentative")
                return
            t0, t1, a, b, r2 = best
            self.ent_t0.delete(0, tk.END)
            self.ent_t0.insert(0, f"{t0:.0f}")
            self.ent_t1.delete(0, tk.END)
            self.ent_t1.insert(0, f"{t1:.0f}")
        else:
            try:
                t0 = float(self.ent_t0.get().strip().replace(",", "."))
                t1 = float(self.ent_t1.get().strip().replace(",", "."))
            except Exception as e:
                self.log.error("COMPUTE_RATE_PARSE", e)
                messagebox.showerror("Vitesse", "Temps début/fin invalides.")
                return

            if t1 <= t0:
                messagebox.showerror("Vitesse", "t fin doit être > t début")
                self.log.warning("COMPUTE_RATE", f"Intervalle invalide t0={t0} | t1={t1}")
                return

            xs = [r.t_s for r in self.rows if t0 <= r.t_s <= t1]
            ys = [r.strain_pct for r in self.rows if t0 <= r.t_s <= t1]

            if len(xs) < 3:
                messagebox.showinfo("Vitesse", "Pas assez de points dans l'intervalle.")
                self.log.warning("COMPUTE_RATE", f"Pas assez de points sur [{t0}; {t1}]")
                return

            a, b, r2 = self._linear_regression(xs, ys)

        self._last_slope_pct_per_s = a
        self._set_rate_display(a)

        try:
            self.plot.set_trendline(a=a, b=b, t0_s=t0, t1_s=t1, r2=r2)
        except Exception as e:
            self.log.error("SET_TRENDLINE", e)

        self.status.set(f"Vitesse calculée sur [{t0:.1f}s ; {t1:.1f}s] — R²={r2:.4f}")
        self.log.info("COMPUTE_RATE", f"t0={t0:.2f} | t1={t1:.2f} | a={a:.8f} %/s | b={b:.8f} | R2={r2:.6f}")

    @staticmethod
    def _linear_regression(x: List[float], y: List[float]) -> Tuple[float, float, float]:
        n = len(x)
        mx = sum(x) / n
        my = sum(y) / n
        sxx = sum((xi - mx) ** 2 for xi in x)
        if sxx == 0:
            return 0.0, my, 0.0
        sxy = sum((xi - mx) * (yi - my) for xi, yi in zip(x, y))
        a = sxy / sxx
        b = my - a * mx
        ss_tot = sum((yi - my) ** 2 for yi in y)
        ss_res = sum((yi - (a * xi + b)) ** 2 for xi, yi in zip(x, y))
        r2 = 0.0 if ss_tot == 0 else max(0.0, 1.0 - ss_res / ss_tot)
        return a, b, r2

    def _on_toggle_auto_rate(self) -> None:
        if self.var_auto_rate.get():
            self.status.set("Mode Auto: recherche d'une zone linéaire représentative (vous pouvez décocher pour choisir t début/fin).")
            self.log.info("AUTO_RATE_MODE", "AUTO")
        else:
            self.status.set("Mode Manuel: choisissez t début et t fin, puis cliquez sur Calculer.")
            self.log.info("AUTO_RATE_MODE", "MANUAL")

    def _auto_select_linear_window(self) -> Optional[Tuple[float, float, float, float, float]]:
        rows = self.rows
        if len(rows) < 10:
            return None

        x = [r.t_s for r in rows]
        y = [r.strain_pct for r in rows]
        n = len(x)
        total_dur = max(1e-9, x[-1] - x[0])

        min_points = max(20, int(0.15 * n))
        min_dur = max(10.0, 0.20 * total_dur)

        Sx = [0.0]
        Sy = [0.0]
        Sxx = [0.0]
        Sxy = [0.0]
        Syy = [0.0]
        for xi, yi in zip(x, y):
            Sx.append(Sx[-1] + xi)
            Sy.append(Sy[-1] + yi)
            Sxx.append(Sxx[-1] + xi * xi)
            Sxy.append(Sxy[-1] + xi * yi)
            Syy.append(Syy[-1] + yi * yi)

        def window_sums(i: int, j: int) -> Tuple[int, float, float, float, float, float]:
            nn = j - i + 1
            sx = Sx[j + 1] - Sx[i]
            sy = Sy[j + 1] - Sy[i]
            sxx = Sxx[j + 1] - Sxx[i]
            sxy = Sxy[j + 1] - Sxy[i]
            syy = Syy[j + 1] - Syy[i]
            return nn, sx, sy, sxx, sxy, syy

        def fit(i: int, j: int) -> Optional[Tuple[float, float, float]]:
            nn, sx, sy, sxx, sxy, syy = window_sums(i, j)
            if nn < 3:
                return None
            mx = sx / nn
            my = sy / nn
            sxx_c = sxx - nn * mx * mx
            if abs(sxx_c) < 1e-12:
                return None
            sxy_c = sxy - nn * mx * my
            a = sxy_c / sxx_c
            b = my - a * mx
            ss_tot = syy - nn * my * my
            ss_res = ss_tot - a * sxy_c
            r2 = 0.0 if ss_tot <= 0 else max(0.0, 1.0 - (ss_res / ss_tot))
            return a, b, r2

        best = None
        best_score = -1.0

        step_i = 5
        step_len = 10
        for i in range(0, n - min_points, step_i):
            max_j = n - 1
            j0 = i + min_points - 1
            if j0 >= n:
                break
            for j in range(j0, max_j + 1, step_len):
                dur = x[j] - x[i]
                if dur < min_dur:
                    continue
                res = fit(i, j)
                if res is None:
                    continue
                a, b, r2 = res

                frac_dur = min(1.0, dur / total_dur)
                frac_pts = min(1.0, (j - i + 1) / n)
                score = r2 * (0.6 * (frac_dur ** 0.5) + 0.4 * (frac_pts ** 0.5))

                if a < 0:
                    continue

                if score > best_score:
                    best_score = score
                    best = (x[i], x[j], a, b, r2)

        if best is None:
            min_points2 = max(30, int(0.10 * n))
            min_dur2 = max(30.0, 0.10 * total_dur)
            for i in range(0, n - min_points2, step_i):
                j0 = i + min_points2 - 1
                if j0 >= n:
                    break
                for j in range(j0, n, step_len):
                    dur = x[j] - x[i]
                    if dur < min_dur2:
                        continue
                    res = fit(i, j)
                    if res is None:
                        continue
                    a, b, r2 = res
                    frac_dur = min(1.0, dur / total_dur)
                    frac_pts = min(1.0, (j - i + 1) / n)
                    score = r2 * (0.6 * (frac_dur ** 0.5) + 0.4 * (frac_pts ** 0.5))
                    if a < 0:
                        continue
                    if score > best_score:
                        best_score = score
                        best = (x[i], x[j], a, b, r2)

        return best

    # ------------- Export -------------
    def _export_csv(self) -> None:
        if self.worker is not None:
            messagebox.showinfo("Action indisponible", "Déconnectez-vous avant d'exporter les données.")
            self.log.warning("EXPORT_CSV_BLOCKED", "Tentative pendant acquisition")
            return

        if not self.rows:
            messagebox.showinfo("Export", "Aucune donnée à exporter.")
            self.log.warning("EXPORT_CSV", "Aucune donnée")
            return

        path = filedialog.asksaveasfilename(
            title="Exporter CSV",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
        )
        if not path:
            self.log.info("EXPORT_CSV_CANCELLED")
            return

        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(["Temps (s)", "Distance (mm)", "Allongement (mm)", "Deformation (%)", "Temperature (°C)"])
                for r in self.rows:
                    w.writerow([r.t_s, r.distance_mm, r.elong_mm, r.strain_pct, r.temp_c])

            self.status.set(f"CSV exporté: {os.path.basename(path)}")
            self.log.file_export("EXPORT_CSV", path)
        except Exception as e:
            self.log.error("EXPORT_CSV", e, f"Path={path}")
            messagebox.showerror("Export CSV", f"Erreur lors de l'export CSV :\n{e}")

    def _export_xlsx(self) -> None:
        if self.worker is not None:
            messagebox.showinfo("Action indisponible", "Déconnectez-vous avant d'exporter les données.")
            self.log.warning("EXPORT_XLSX_BLOCKED", "Tentative pendant acquisition")
            return

        if not self.rows:
            messagebox.showinfo("Export", "Aucune donnée à exporter.")
            self.log.warning("EXPORT_XLSX", "Aucune donnée")
            return

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except Exception as e:
            self.log.error("EXPORT_XLSX_IMPORT", e)
            messagebox.showerror("Export", "openpyxl manquant: pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            title="Exporter Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            self.log.info("EXPORT_XLSX_CANCELLED")
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Mesures"

            headers = ["Temps (s)", "Distance (mm)", "Allongement (mm)", "Deformation (%)", "Temperature (°C)"]
            ws.append(headers)
            for r in self.rows:
                ws.append([r.t_s, r.distance_mm, r.elong_mm, r.strain_pct, r.temp_c])

            for i, h in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

            wb.save(path)
            self.status.set(f"Excel exporté: {os.path.basename(path)}")
            self.log.file_export("EXPORT_XLSX", path)
        except Exception as e:
            self.log.error("EXPORT_XLSX", e, f"Path={path}")
            messagebox.showerror("Export Excel", f"Erreur lors de l'export Excel :\n{e}")

    # ------------- Report (PDF) -------------
    def _generate_report(self) -> None:
        if self.worker is not None:
            messagebox.showinfo("Action indisponible", "Déconnectez-vous avant de générer un rapport.")
            self.log.warning("GENERATE_REPORT_BLOCKED", "Tentative pendant acquisition")
            return

        if not self.rows:
            messagebox.showinfo("Rapport", "Aucune donnée.")
            self.log.warning("GENERATE_REPORT", "Aucune donnée")
            return

        out_dir = filedialog.askdirectory(title="Dossier de sortie")
        if not out_dir:
            self.log.info("GENERATE_REPORT_CANCELLED", "Aucun dossier choisi")
            return

        meta = self._ask_meta()
        if meta is None:
            self.log.info("GENERATE_REPORT_CANCELLED", "Fenêtre méta annulée")
            return

        try:
            t_tot = self.rows[-1].t_s
            t_val, t_unit = self._time_display(t_tot)
            t_tot_disp = f"{t_val:.2f} {t_unit}"
            eps_max = max(r.strain_pct for r in self.rows)
            temp_mean = sum(r.temp_c for r in self.rows) / len(self.rows)

            rate_txt = self.var_rate.get()
            rate = rate_txt if rate_txt != "—" else "(non calculé)"

            x, strain, temp = self.plot.get_arrays()
            graph_def = os.path.join(out_dir, "graph_deformation.png")
            graph_temp = os.path.join(out_dir, "graph_temperature.png")
            factor, label = self.plot.get_time_display_params()

            trend = getattr(self.plot, "_trend", None)

            results = {
                "Temps total": t_tot_disp,
                "Déformation max (%)": f"{eps_max:.4f}",
                "Vitesse de déformation": str(rate),
                "Température moyenne (°C)": f"{temp_mean:.2f}",
            }

            regression = None
            if trend:
                a = float(trend["a"])
                b = float(trend["b"])
                t0 = float(trend["t0"])
                t1 = float(trend["t1"])
                r2 = trend.get("r2", None)

                regression = {
                    "t0 (s)": t0,
                    "t1 (s)": t1,
                    "a (%/s)": a,
                    "b (%)": b,
                    "R²": None if r2 is None else float(r2),
                    "Equation (s)": f"y = {a:.6g}·t + {b:.6g} (t en s)",
                }

            self._save_graphs(
                x, strain, temp,
                graph_def, graph_temp,
                time_factor=factor, time_label=label,
                trend=trend
            )

            pdf_path = os.path.join(out_dir, f"Rapport_Fluage_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")

            logo_iut_path = resource_path("assets", "logoIutEvreux.png")
            if not os.path.exists(logo_iut_path):
                logo_iut_path = None

            generate_pdf_report(
                output_pdf=pdf_path,
                app_version=APP_VERSION,
                meta={
                    "Lancé par": meta["launchBy"],
                    "Capteur": meta["sensor"],
                    "Matériau": meta["material"],
                    "Longueur initiale (mm)": meta["initial_length"],
                    "Unité temps (affichage)": t_unit,
                    "Session ID": self.log.session_id,
                    "Machine": self.log.machine_name,
                    "IP machine": self.log.machine_ip,
                    "Version logicielle": APP_VERSION,
                },
                results=results,
                regression=regression,
                LOGO_IUT_PATH=logo_iut_path,
                graph_def_path=graph_def,
                graph_temp_path=graph_temp,
            )

            self.log.file_export("EXPORT_GRAPH_DEFORMATION", graph_def)
            self.log.file_export("EXPORT_GRAPH_TEMPERATURE", graph_temp)
            self.log.file_export("GENERATE_PDF", pdf_path)

            messagebox.showinfo("Rapport", f"PDF généré:\n{pdf_path}")
            self.status.set("Rapport PDF généré")
        except Exception as e:
            self.log.error("GENERATE_REPORT", e)
            messagebox.showerror("Rapport", f"Erreur génération PDF:\n{e}")

    def _ask_meta(self) -> Optional[dict]:
        dlg = tb.Toplevel(self.root)
        dlg.title("Infos échantillon")
        dlg.geometry("720x320")
        dlg.resizable(False, False)

        res: dict = {}

        def add_row(r, label, default=""):
            tb.Label(frm, text=label).grid(row=r, column=0, sticky="w", pady=6)
            ent = tb.Entry(frm)
            ent.insert(0, default)
            ent.grid(row=r, column=1, sticky="ew", pady=6)
            return ent

        frm = tb.Frame(dlg, padding=14)
        frm.pack(fill=tk.BOTH, expand=True)
        frm.columnconfigure(1, weight=1)

        ent_user = add_row(0, "Lancé par", get_current_username())
        ent_sensor = add_row(1, "Capteurs utilisés", "HC-SR04 + NTC")
        ent_mat = add_row(2, "Matériau", "")
        ent_init = add_row(3, "Longueur initiale (mm)", "85")

        btn_row = tb.Frame(frm)
        btn_row.grid(row=6, column=0, columnspan=2, sticky="e", pady=(18, 0))

        ok = {"v": False}

        def on_ok():
            res["launchBy"] = ent_user.get().strip() or "—"
            res["sensor"] = ent_sensor.get().strip() or "—"
            res["material"] = ent_mat.get().strip() or "—"
            res["initial_length"] = ent_init.get().strip() or "—"
            ok["v"] = True
            dlg.destroy()

        def on_cancel():
            dlg.destroy()

        tb.Button(btn_row, text="Annuler", bootstyle="secondary", command=on_cancel).pack(side=tk.RIGHT)
        tb.Button(btn_row, text="OK", bootstyle="success", command=on_ok).pack(side=tk.RIGHT, padx=8)

        dlg.grab_set()
        self.root.wait_window(dlg)
        return res if ok["v"] else None

    @staticmethod
    def _save_graphs(
        x, strain, temp, out_def, out_temp,
        time_factor: float = 1.0, time_label: str = "s",
        trend: dict | None = None
    ):
        import matplotlib.pyplot as plt

        plt.figure()
        x_view = [v / time_factor for v in x]
        plt.plot(x_view, strain)

        if trend is not None:
            try:
                a = float(trend.get("a"))
                b = float(trend.get("b"))
                t0 = float(trend.get("t0"))
                t1 = float(trend.get("t1"))
                r2 = trend.get("r2", None)

                x_seg = [t0 / time_factor, t1 / time_factor]
                y_seg = [a * t0 + b, a * t1 + b]

                plt.plot(x_seg, y_seg, linestyle="--", color="red", linewidth=1.5)

                a_disp = a * time_factor
                eq = f"y = {a_disp:.6g} x + {b:.6g}"
                if r2 is not None:
                    eq += f"\nR² = {float(r2):.4f}"

                ax = plt.gca()
                ax.text(
                    0.98, 0.02, eq,
                    transform=ax.transAxes,
                    ha="right", va="bottom",
                    fontsize=12, color="red"
                )
            except Exception:
                pass

        plt.xlabel(f"Temps ({time_label})")
        plt.ylabel("Déformation (%)")
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(out_def, dpi=160)
        plt.close()

        plt.figure()
        x_view = [v / time_factor for v in x]
        plt.plot(x_view, temp)
        plt.xlabel(f"Temps ({time_label})")
        plt.ylabel("Température (°C)")
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(out_temp, dpi=160)
        plt.close()

    def run(self) -> None:
        try:
            self.root.mainloop()
        except Exception as e:
            self.log.error("MAINLOOP", e)
            raise


# ==================================================================================================
# PDF REPORT
# ==================================================================================================

def generate_pdf_report(
    output_pdf: str,
    app_version: str,
    meta: dict,
    results: dict,
    regression: dict | None,
    LOGO_IUT_PATH: str | None,
    graph_def_path: str,
    graph_temp_path: str,
) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate,
        PageBreak,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        Image,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    doc = SimpleDocTemplate(
        output_pdf,
        pagesize=A4,
        leftMargin=2.0 * cm,
        rightMargin=2.0 * cm,
        topMargin=1.6 * cm,
        bottomMargin=1.6 * cm,
        title="Rapport d'analyse - Fluage",
        author=str(meta.get("Lancé par", "")),
        subject="Rapport généré le " + datetime.now().strftime("%d/%m/%Y à %H:%M:%S") + " via FluageAutomation v" + app_version,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleCenter", parent=styles["Title"], alignment=TA_CENTER, spaceAfter=10))
    styles.add(ParagraphStyle(name="H2", parent=styles["Heading2"], spaceBefore=12, spaceAfter=6))
    styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=9, leading=11))
    styles.add(ParagraphStyle(name="Caption", parent=styles["BodyText"], fontSize=9, leading=11, alignment=TA_CENTER, textColor=colors.grey))

    story = []

    subtitle = Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')} - Version utilisée : {app_version}",
        styles["Small"]
    )

    header_cells = []
    if LOGO_IUT_PATH and os.path.exists(LOGO_IUT_PATH):
        try:
            from reportlab.lib.utils import ImageReader
            img_reader = ImageReader(LOGO_IUT_PATH)
            iw, ih = img_reader.getSize()
            max_w, max_h = 3.2 * cm, 2.2 * cm
            scale = min(max_w / float(iw), max_h / float(ih))
            header_cells.append(Image(LOGO_IUT_PATH, width=iw * scale, height=ih * scale))
        except Exception:
            header_cells.append("")
    else:
        header_cells.append("")

    header_cells.append(Paragraph("<b>Rapport d'analyse - Essai de fluage</b><br/><font size=10>IUT d'Évreux</font>", styles["Title"]))
    header = Table([header_cells], colWidths=[2.6 * cm, 14.4 * cm])
    header.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.black),
            ]
        )
    )
    story.append(header)
    story.append(Spacer(1, 6))
    story.append(subtitle)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Informations", styles["H2"]))
    meta_rows = [["Champ", "Valeur"]] + [[k, str(v)] for k, v in meta.items()]
    t = Table(meta_rows, colWidths=[9.5 * cm, 7.0 * cm])
    t.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9ecef")),
                ("LINEABOVE", (0, 0), (-1, 0), 0.8, colors.black),
                ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.black),
                ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.grey),
            ]
        ),
    )
    story.append(t)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Résultats clés", styles["H2"]))
    res_rows = [["Indicateur", "Valeur"]] + [[k, str(v)] for k, v in results.items()]
    t2 = Table(res_rows, colWidths=[9.5 * cm, 7.0 * cm])
    t2.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9ecef")),
                ("LINEABOVE", (0, 0), (-1, 0), 0.8, colors.black),
                ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.black),
                ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.grey),
            ]
        ),
    )
    story.append(t2)
    story.append(Spacer(1, 14))

    reg = regression
    if isinstance(reg, dict):
        story.append(Paragraph("Régression linéaire (zone sélectionnée)", styles["H2"]))

        data = [
            ["Paramètre", "Valeur"],
            [Paragraph("t<sub>0</sub> (s)", styles["BodyText"]), f"{float(reg.get('t0 (s)', 0.0)):.2f}"],
            [Paragraph("t<sub>1</sub> (s)", styles["BodyText"]), f"{float(reg.get('t1 (s)', 0.0)):.2f}"],
            ["a (%/s)", f"{float(reg.get('a (%/s)', 0.0)):.6g}"],
            ["b (%)", f"{float(reg.get('b (%)', 0.0)):.6g}"],
        ]
        r2v = reg.get("R²", None)
        data.append(["R² (coefficient de détermination)", "—" if r2v is None else f"{float(r2v):.4f}"])

        t_reg = Table(data, colWidths=[9.5 * cm, 7.0 * cm])
        t_reg.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9ecef")),
                    ("LINEABOVE", (0, 0), (-1, 0), 0.8, colors.black),
                    ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.black),
                    ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.grey),
                ]
            )
        )
        story.append(t_reg)

        eq = reg.get("Equation (s)", None)
        if eq:
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"<b>Équation :</b> {eq}", styles["BodyText"]))

        story.append(Spacer(1, 14))

    story.append(PageBreak())
    story.append(Paragraph("Courbes", styles["H2"]))
    story.append(Spacer(1, 8))

    if graph_def_path and os.path.exists(graph_def_path):
        story.append(Image(graph_def_path, width=15 * cm, height=11.25 * cm))
        story.append(Paragraph("Figure 1 - Déformation en fonction du temps", styles["Caption"]))
        story.append(Spacer(1, 10))
    else:
        story.append(Paragraph("Graphique déformation introuvable.", styles["Small"]))

    if graph_temp_path and os.path.exists(graph_temp_path):
        story.append(Image(graph_temp_path, width=15 * cm, height=11.25 * cm))
        story.append(Paragraph("Figure 2 - Température en fonction du temps", styles["Caption"]))
        story.append(Spacer(1, 6))
    else:
        story.append(Paragraph("Graphique température introuvable.", styles["Small"]))

    def _on_page(canvas, _doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(A4[0] - 2.0 * cm, 1.0 * cm, f"Page {_doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)


def main() -> None:
    app = FluageApp()
    app.run()


if __name__ == "__main__":
    main()